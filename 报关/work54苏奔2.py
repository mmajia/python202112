from selenium import webdriver
import openpyxl
import re

#Chrome
from selenium.webdriver import ActionChains, Keys
from selenium.webdriver.common.by import By



#excel
wb=openpyxl.load_workbook(r'C:\Users\MaJia\Desktop\2月11日(1).xlsx',data_only=True)#data_only=True 直接取公式之后的结果
sheet=wb['954精密']
#excel
a = int(input('请输入首行号'))
b = int(input('请输入结尾行号'))
astart=a
aend=b
print(astart,aend)
totalCarton = 0#总包装箱
totalGrossWeight = 0#总毛重
totalProductNetW = 0#总产品净重
#excel
while astart<=aend:
    #总包装箱
    Carton=int(sheet.cell(astart,10).value)
    totalCarton +=Carton
    #总毛重
    GrossWeight = int(sheet.cell(astart,13).value)
    totalGrossWeight += GrossWeight
    #总产品净重
    ProductNetW = int(sheet.cell(astart,9).value)
    totalProductNetW += ProductNetW
    astart += 1
astart=a
aend=b


#'''
print('总包装箱'+str(totalCarton))
print('总毛重' + str(totalGrossWeight))
print('总产品净重' + str(totalProductNetW))
#'''

#54苏奔
driver = webdriver.Chrome()
driver.get('http://61.177.55.210/mrs954/')
#登录页面
driver.find_element(By.ID,"edtUserID").send_keys('drwang')
driver.find_element(By.ID,"edtPassword").send_keys('456852bt')
driver.find_element(By.ID,"btnLogin").click()
#登陆成功选择表格
menu1_1=driver.find_element(By.XPATH,'//*[@id="Menu1_1"]/nobr')
ActionChains(driver).move_to_element(menu1_1).perform()
Menu1_4=driver.find_element(By.XPATH,'//*[@id="Menu1_4"]')
#driver.find_elements(By.XPATH,'//*[@id="Menu1_4"]').click()
ActionChains(driver).move_to_element(Menu1_4).perform()
driver.find_element(By.ID,'Menu1_4').click()
#开始输入
driver.switch_to.frame('iframe1')#切换frame正确用法
driver.find_element(By.ID,"edtPackageCount").send_keys(totalCarton)#总件数
driver.find_element(By.ID,"edtGrossWT").send_keys(totalGrossWeight)#总毛重
driver.find_element(By.ID,"edtNetWT").send_keys(totalProductNetW)#总净重

#每个料循环输入
while astart<=aend:
    #料号
    edtMaterial=sheet.cell(astart,3).value
    print(edtMaterial)
    driver.find_element(By.ID, "edtMaterial").send_keys(edtMaterial)  # 料号
    driver.find_element(By.ID, "edtMaterial").send_keys(Keys.TAB)  # 按tab键
    #备注
    remarks = sheet.cell(astart, 17).value
    if remarks== None:
        remarks=''
    # 公式
    formula=sheet.cell(astart, 14).value#公式
    searchObj = re.search(r'^(\d+)(?=\*)', formula)  # 匹配第一个数字，包装箱
    #print(searchObj.group(1))
    searchObj2 = re.search(r'(?<=\*)(\d+)', formula)  # 匹配第2个数字，包装箱的数量
    #print(searchObj2.group(1))
    driver.find_element(By.ID, "edtCaseQty").send_keys(searchObj.group(1))  # 多少个包装
    driver.find_element(By.ID, "edtCasePerQty").send_keys(searchObj2.group(1))  # 多少个包装的数量
    searchObj3 = re.search(r'\d+\*\d+\+(.*)', formula)  # 匹配第1个算式之后的其他所有算式

    if searchObj3 == None:
        driver.find_element(By.ID, "edtRemark").send_keys(remarks)  # 备注
        driver.find_element(By.ID, 'btnSave').click()  # 保存当前料号
    else:
        driver.find_element(By.ID, "edtRestQty").send_keys(Keys.BACK_SPACE)  # 删除键，删掉自带的0
        driver.find_element(By.ID, "edtRestQty").send_keys(eval(searchObj3.group(1)))  # 零箱
        driver.find_element(By.ID, "edtRemark").send_keys(searchObj3.group(1)+'/'+remarks)  # 备注
        driver.find_element(By.ID, 'btnSave').click()  # 保存当前料号
        #print(searchObj3.group(1))
        #print(eval(searchObj3.group(1)))
    astart+=1
